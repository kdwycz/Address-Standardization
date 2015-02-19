#!/usr/bin/env python
# coding=utf-8

__author__ = 'kdwycz'

import os
from openpyxl import load_workbook
from global_list import *
import baidu.map as baidu
import platform

def Analyse(address):
    """分析地区名是否合法函数：
    输入：地址字符串,百度APIKEY
    返回：一个元组，第一个元素是True/False。如果为True，
    第二个元素是一个字典。表示地区的省市县
    """
    flag = False
    for divisions in DIVISIONS:
        if divisions in address:
            flag = True
            break
    ddict = {}
    if flag:
        global NUM
        xy = baidu.findxy(address, KEY[NUM % APINUM])
        NUM += 1
        if xy == (0,0):
            flag = False
        else:
            ddict = baidu.finddir(xy, KEY[NUM % APINUM])
            if ddict[u'province'] == u'error':
                flag = False
            NUM += 1
    return (flag,ddict)

def CreateDict(sheetname,maindict):
    """生成基础字典供之后分析函数：
    输入：表名，字典名
    返回：在此过程中发生错误个数
    """
    error = 0
    ignore = 0
    inws = inwb.get_sheet_by_name(sheetname)
    sheetrow = inws.get_highest_row()
    sheetcol = inws.get_highest_column()
    for i in range(2,sheetrow+1):
        num = inws.cell(row=i,column=1).value
        address = inws.cell(row=i,column=2).value
        flag = True
        if num and address:
            addr = inws.cell(row=i,column=sheetcol).value
            if addr == None or addr == 'Error':
                ddict = Analyse(address)
                if ddict[0]:
                    ddict = ddict[1]
                    maindict[num] = ddict
                    addr = ddict['province'] + ' ' + ddict['city'] + ' ' + ddict['district']
                    inws.cell(row=i,column=sheetcol).value = addr
                else:
                    flag = False
            else:
                ddict = {'province' : addr.split()[0], 'city' : addr.split()[1], 'district' : addr.split()[1]}
                maindict[num] = ddict
                ignore += 1
        else:
            flag = False
        if not flag:
            inws.cell(row=i,column=sheetcol).value = 'Error'
            error += 1
    return (error,ignore)

def CheckDict(sheetname,maindict):
    """读取表和与主字典对比：
    输入：表名，字典名
    返回：在此过程中发生错误个数
    """
    error = 0
    ignore = 0
    inws = inwb.get_sheet_by_name(sheetname)
    sheetrow = inws.get_highest_row()
    sheetcol = inws.get_highest_column()
    for i in range(2,sheetrow+1):
        num = inws.cell(row=i,column=1).value
        for j in xrange(2,sheetcol/2+2):
            k = sheetcol/2
            address = inws.cell(row=i,column=j).value
            flag = True
            if num and address:
                if sheetname == 'LH':
                    address += u'市'
                addr = inws.cell(row=i,column=j+k).value
                if addr == None or addr == 'Error':
                    ddict = Analyse(address)
                    if ddict[0]:
                        ddict = ddict[1]
                        if num not in maindict:
                            flag = False
                        elif ddict['province'] != maindict[num]['province']:
                            result = 4
                        elif ddict['city'] != maindict[num]['city']:
                            result = 3
                        elif ddict['district'] != maindict[num]['district']:
                            result = 2
                        else:
                            result = 1
                    else:
                        flag = False
                else:
                    ignore += 1
                    result = addr
            else:
                flag = False
            if not flag:
                result = 'Error'
                error += 1
            inws.cell(row=i,column=j+k).value = result
    return (error,ignore)

def CheckDoubleDict():
    """读取两个表对比(LG表)：
    返回：在此过程中发生错误个数
    """
    error = 0
    ignore = 0
    inws = inwb.get_sheet_by_name(MIXSET)
    sheetrow = inws.get_highest_row()
    sheetcol = inws.get_highest_column()
    for i in range(2,sheetrow+1):
        num1 = inws.cell(row=i,column=1).value
        num2 = inws.cell(row=i,column=2).value
        flag = True
        addr = inws.cell(row=i,column=sheetcol).value
        if addr == None or addr == 'Error':
            if num1 and num2 and (num1 in rootdict) and (num2 in hoteldict):
                if rootdict[num1]['province'] != hoteldict[num2]['province']:
                    result = 4
                elif rootdict[num1]['city'] != hoteldict[num2]['city']:
                    result = 3
                elif rootdict[num1]['district'] != hoteldict[num2]['district']:
                    result = 2
                else:
                    result = 1
            else:
                flag = False
            if not flag:
                result = 'Error'
                error += 1
            inws.cell(row=i,column=sheetcol).value = result
        else:
            ignore += 1
    return (error,ignore)

if __name__ == '__main__':

    print(u'程序开始运行...')

#    sysstr = platform.system()
#    if sysstr == "Windows":
#        OS = 1

    if not os.path.exists('output'):
        os.mkdir('output')

    try:
        inwb = load_workbook('./input/data.xlsx')
    except:
        print (u'文件读取错误')
    else:
	    rootdict = {}
	    hoteldict = {}

	    report = CreateDict(ROOTNAME,rootdict)
	    print (u"读取主地址并进行标准化输出:出现错误数: %d,忽略已处理数据数: %d" %report)

	    report = CreateDict(HOTELNAME,hoteldict)
	    print (u"读取旅馆地址并进行标准化输出:出现错误数: %d,忽略已处理数据数: %d" %report)

	    for item in ROOTSET:
	        report = CheckDict(item,rootdict)
	        print (u"读取%s表并进行比对:出现错误数: %d,忽略已处理数据数: %d" %((item,)+report))

	    report = CheckDoubleDict()
	    print (u"读取%s表并进行比对:出现错误数: %d,忽略已处理数据数: %d" %((MIXSET,)+report))

	    inwb.save('./output/tst.xlsx')
	    print (u'全部完成')
